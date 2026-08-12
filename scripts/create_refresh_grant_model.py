from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.comments import Comment
from openpyxl.worksheet.dimensions import ColumnDimension

OUTPUT = Path("/home/ubuntu/AI-Code-Reviewer/refresh_grant_model_template.xlsx")

DARK_GREEN = "135B44"
LIGHT_GREEN = "CFE9E0"
LIGHT_GRAY = "E7E5E4"
WHITE = "FFFFFF"
BLUE = "0000FF"
BLACK = "000000"
GREEN = "008000"
PURPLE = "800080"

thin_gray = Side(style="thin", color="A6A6A6")
medium_green = Side(style="medium", color=DARK_GREEN)
double_black = Side(style="double", color="000000")


def styled_title(ws, title, subtitle, unit_note):
    ws["C3"] = title
    ws["C3"].fill = PatternFill("solid", fgColor=DARK_GREEN)
    ws["C3"].font = Font(name="Arial", size=16, bold=True, color=WHITE)
    ws["C3"].alignment = Alignment(horizontal="left")
    ws.merge_cells("C3:H3")
    ws["C5"] = subtitle
    ws["C5"].font = Font(name="Arial", size=11, bold=True, color=BLACK)
    ws["C6"] = unit_note
    ws["C6"].font = Font(name="Arial", size=10, italic=True, color=BLACK)
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 20
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.oddFooter.center.text = "&A | Page &P"


def section(ws, row, title):
    ws[f"C{row}"] = title
    ws[f"C{row}"].fill = PatternFill("solid", fgColor=LIGHT_GREEN)
    ws[f"C{row}"].font = Font(name="Arial", bold=True, color=BLACK)
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=8)


def set_input(ws, row, label, value=None, number_format="General", note="Management policy input — enter only Board/CFO/HR-approved data."):
    ws[f"C{row}"] = label
    ws[f"D{row}"] = value
    ws[f"D{row}"].font = Font(name="Arial", color=BLUE)
    ws[f"D{row}"].number_format = number_format
    ws[f"D{row}"].alignment = Alignment(horizontal="right")
    ws[f"D{row}"].comment = Comment(note, "Manus AI")


def set_link(ws, row, label, formula, number_format="General"):
    ws[f"C{row}"] = label
    ws[f"D{row}"] = formula
    ws[f"D{row}"].font = Font(name="Arial", color=GREEN)
    ws[f"D{row}"].number_format = number_format
    ws[f"D{row}"].alignment = Alignment(horizontal="right")


def set_formula(ws, row, label, formula, number_format="General", bold=False):
    ws[f"C{row}"] = label
    ws[f"D{row}"] = formula
    ws[f"D{row}"].font = Font(name="Arial", color=BLACK, bold=bold)
    ws[f"D{row}"].number_format = number_format
    ws[f"D{row}"].alignment = Alignment(horizontal="right")
    if bold:
        ws[f"C{row}"].font = Font(name="Arial", bold=True, color=BLACK)
        ws[f"C{row}"].border = Border(top=medium_green)
        ws[f"D{row}"].border = Border(top=medium_green)


def auto_fit(ws):
    for col in range(3, ws.max_column + 1):
        letter = ws.cell(row=1, column=col).column_letter
        length = 10
        for row in range(1, ws.max_row + 1):
            value = ws.cell(row=row, column=col).value
            if value is not None:
                length = max(length, min(len(str(value)) + 2, 58))
        ws.column_dimensions[letter].width = length


def add_notes(ws, start_row, text):
    ws[f"C{start_row}"] = "مدل فقط پس از ورود داده‌های تاییدشده خروجی عددی ایجاد می‌کند."
    ws[f"C{start_row}"].font = Font(name="Arial", italic=True, color=BLACK)
    ws.merge_cells(start_row=start_row, start_column=3, end_row=start_row, end_column=8)
    ws[f"C{start_row + 1}"] = text
    ws[f"C{start_row + 1}"].font = Font(name="Arial", italic=True, color=BLACK)
    ws.merge_cells(start_row=start_row + 1, start_column=3, end_row=start_row + 1, end_column=8)


wb = Workbook()
inputs = wb.active
inputs.title = "Inputs"
model = wb.create_sheet("Model")
board = wb.create_sheet("Board Summary")

# Inputs tab
styled_title(inputs, "Refresh Grant Model — Inputs", "Management inputs and approved policy multipliers", "(USD, shares/units, months and multiples as labeled)")
section(inputs, 8, "Individual and plan inputs — enter approved data in blue cells")
inputs["C9"] = "Input"
inputs["D9"] = "Value"
inputs["E9"] = "Definition / Source"
for cell in inputs[9][2:5]:
    cell.fill = PatternFill("solid", fgColor=LIGHT_GRAY)
    cell.font = Font(name="Arial", bold=True, color=BLACK)

set_input(inputs, 10, "Annual target equity value for role (TEV)", None, "$#,##0.0;($#,##0.0);-", "Source: Board-approved compensation framework or HR market review; as-of date required.")
inputs["E10"] = "Board-approved annual target equity value for the employee's role/level."
set_input(inputs, 11, "Target coverage months (CM)", None, "0", "Source: Board-approved retention policy; typically expressed in months.")
inputs["E11"] = "Months of future unvested-equity coverage targeted by policy."
set_input(inputs, 12, "Fair market value per share / unit (FMV)", None, "0.00", "Source: Latest approved valuation / 409A-equivalent assessment; date and approver required.")
inputs["E12"] = "Confirmed fair market value per share/unit; do not use enterprise value."
set_input(inputs, 13, "Current unvested options / units (UO)", None, "#,##0", "Source: Cap table administrator / equity platform; as-of grant review date.")
inputs["E13"] = "Granted and unvested options/units held by the employee."
set_input(inputs, 14, "Individual grant cap (IC)", None, "#,##0", "Source: Compensation committee policy; annual or cycle-specific cap.")
inputs["E14"] = "Maximum refresh shares/units permitted for one employee in this cycle."
set_input(inputs, 15, "Remaining ESOP pool capacity (PC)", None, "#,##0", "Source: Cap table / board-approved option pool; as-of grant review date.")
inputs["E15"] = "Unallocated shares/units still available in the ESOP pool."
set_input(inputs, 16, "Basic shares / units outstanding", None, "#,##0", "Source: Cap table; required only for ownership-dilution reporting.")
inputs["E16"] = "Basic shares/units outstanding before the proposed refresh grant."

section(inputs, 19, "Policy multipliers — initial values are proposed policy settings; Board approval required")
inputs["C20"] = "Multiplier"
inputs["D20"] = "Proposed value"
inputs["E20"] = "Use case"
for cell in inputs[20][2:5]:
    cell.fill = PatternFill("solid", fgColor=LIGHT_GRAY)
    cell.font = Font(name="Arial", bold=True, color=BLACK)

policy_values = [
    (21, "Performance factor — outstanding", 1.25, "Use only after calibrated annual performance review."),
    (22, "Performance factor — strong", 1.00, "Standard high-performance outcome."),
    (23, "Performance factor — meets expectations", 0.75, "Minimum retention coverage subject to role needs."),
    (24, "Performance factor — needs improvement", 0.00, "No automatic refresh grant; exception needs committee approval."),
    (25, "Criticality factor — critical role", 1.20, "Security, core AI, architecture or hard-to-replace role."),
    (26, "Criticality factor — standard role", 1.00, "Default role-criticality setting."),
    (27, "Criticality factor — lower urgency", 0.85, "Use only with documented business rationale."),
    (28, "Retention risk factor — documented high risk", 1.15, "Stay interview and mitigation plan required."),
    (29, "Retention risk factor — normal", 1.00, "Default risk setting."),
    (30, "Retention risk factor — low", 0.90, "Use for budget prioritization; not a punitive factor."),
]
for row, label, value, description in policy_values:
    set_input(inputs, row, label, value, "0.00x", "Source: Internal policy recommendation in REFRESH_GRANT_MODEL.md; requires Compensation Committee approval before operational use.")
    inputs[f"E{row}"] = description

add_notes(inputs, 33, "All blue-font cells are hardcoded inputs or policy parameters and include audit comments.")
inputs.print_area = "B2:H35"

# Model tab
styled_title(model, "Refresh Grant Model — Calculation", "Formula-driven calculation; cross-sheet references are displayed in green", "(USD, shares/units and multiples as labeled)")
section(model, 8, "Linked inputs")
for idx, (label, source_cell, fmt) in enumerate([
    ("Annual target equity value (TEV)", "Inputs!D10", "$#,##0.0;($#,##0.0);-"),
    ("Target coverage months (CM)", "Inputs!D11", "0"),
    ("Fair market value per share / unit (FMV)", "Inputs!D12", "0.00"),
    ("Current unvested options / units (UO)", "Inputs!D13", "#,##0"),
    ("Individual grant cap (IC)", "Inputs!D14", "#,##0"),
    ("Remaining ESOP pool capacity (PC)", "Inputs!D15", "#,##0"),
    ("Basic shares / units outstanding", "Inputs!D16", "#,##0"),
], start=10):
    set_link(model, idx, label, f"={source_cell}", fmt)

section(model, 19, "Selected multipliers — set by committee after calibration")
model["C20"] = "Performance factor (PF)"
model["D20"] = "=IFERROR(INDEX(Inputs!$D$21:$D$24,MATCH(E20,Inputs!$C$21:$C$24,0)),\"\")"
model["D20"].font = Font(name="Arial", color=GREEN)
model["E20"] = "Performance factor — strong"
model["E20"].font = Font(name="Arial", color=BLUE)
model["E20"].comment = Comment("Source: Compensation committee calibrated annual review. Select label exactly as shown in Inputs!C21:C24.", "Manus AI")
model["C21"] = "Criticality factor (CF)"
model["D21"] = "=IFERROR(INDEX(Inputs!$D$25:$D$27,MATCH(E21,Inputs!$C$25:$C$27,0)),\"\")"
model["D21"].font = Font(name="Arial", color=GREEN)
model["E21"] = "Criticality factor — standard role"
model["E21"].font = Font(name="Arial", color=BLUE)
model["E21"].comment = Comment("Source: CTO/CEO role-criticality review with documented business rationale.", "Manus AI")
model["C22"] = "Retention risk factor (RF)"
model["D22"] = "=IFERROR(INDEX(Inputs!$D$28:$D$30,MATCH(E22,Inputs!$C$28:$C$30,0)),\"\")"
model["D22"].font = Font(name="Arial", color=GREEN)
model["E22"] = "Retention risk factor — normal"
model["E22"].font = Font(name="Arial", color=BLUE)
model["E22"].comment = Comment("Source: HR/manager stay interview and retention-risk review; selection must be documented.", "Manus AI")
for r in [20, 21, 22]:
    model[f"D{r}"].number_format = "0.00x"

section(model, 25, "Refresh Grant calculation")
set_formula(model, 26, "Target outstanding value (TOV)", '=IF(OR(NOT(ISNUMBER(D10)),NOT(ISNUMBER(D11))),"",D10*(D11/12))', "$#,##0.0;($#,##0.0);-")
set_formula(model, 27, "Current unvested reference value (CURV)", '=IF(OR(NOT(ISNUMBER(D12)),NOT(ISNUMBER(D13))),"",D12*D13)', "$#,##0.0;($#,##0.0);-")
set_formula(model, 28, "Pre-adjustment retention gap (PARG)", '=IF(OR(NOT(ISNUMBER(D26)),NOT(ISNUMBER(D27))),"",MAX(D26-D27,0))', "$#,##0.0;($#,##0.0);-")
set_formula(model, 29, "Gross refresh shares / units (GRS)", '=IF(OR(NOT(ISNUMBER(D28)),NOT(ISNUMBER(D12)),D12<=0,NOT(ISNUMBER(D20)),NOT(ISNUMBER(D21)),NOT(ISNUMBER(D22))),"",(D28/D12)*D20*D21*D22)', "#,##0")
set_formula(model, 30, "Recommended refresh grant (RRG)", '=IF(OR(NOT(ISNUMBER(D29)),NOT(ISNUMBER(D14)),NOT(ISNUMBER(D15))),"",MIN(D29,D14,D15))', "#,##0", bold=True)
set_formula(model, 31, "Reference value of recommended grant", '=IF(OR(NOT(ISNUMBER(D30)),NOT(ISNUMBER(D12))),"",D30*D12)', "$#,##0.0;($#,##0.0);-")
set_formula(model, 32, "Implied basic-share dilution", '=IF(OR(NOT(ISNUMBER(D30)),NOT(ISNUMBER(D16)),D16<=0),"",D30/D16)', "#,##0.0%")
set_formula(model, 33, "Capacity remaining after recommendation", '=IF(OR(NOT(ISNUMBER(D15)),NOT(ISNUMBER(D30))),"",D15-D30)', "#,##0")
set_formula(model, 34, "Review status", '=IF(OR(NOT(ISNUMBER(D12)),D12<=0,NOT(ISNUMBER(D14)),NOT(ISNUMBER(D15))),"INPUTS REQUIRED",IF(D30>D15,"POOL CAP BREACH",IF(D30>D14,"INDIVIDUAL CAP BREACH","READY FOR COMMITTEE REVIEW")))')
model["D34"].font = Font(name="Arial", color=BLACK, bold=True)

add_notes(model, 37, "Formula logic: RRG = MIN(((MAX((TEV × CM/12) − (UO × FMV), 0) / FMV) × PF × CF × RF), IC, PC).")
model.print_area = "B2:H39"

# Board Summary tab
styled_title(board, "Refresh Grant — Board Summary", "Aggregate control view for Compensation Committee and Board", "(Template; no employee data is pre-populated)")
section(board, 8, "Cycle-level controls")
board["C9"] = "Metric"
board["D9"] = "Formula / result"
board["E9"] = "Board interpretation"
for cell in board[9][2:5]:
    cell.fill = PatternFill("solid", fgColor=LIGHT_GRAY)
    cell.font = Font(name="Arial", bold=True, color=BLACK)

metrics = [
    (10, "Proposed refresh shares / units", "=Model!D30", "Single-employee output; aggregate separately for batch awards." , "#,##0"),
    (11, "Reference value of grant", "=Model!D31", "Reference value based on current FMV; not a promise of future value.", "$#,##0.0;($#,##0.0);-"),
    (12, "ESOP pool capacity before grant", "=Inputs!D15", "Confirm against cap table immediately before Board approval.", "#,##0"),
    (13, "ESOP pool capacity after grant", "=Model!D33", "Must remain non-negative; Board approval required for any pool expansion.", "#,##0"),
    (14, "Implied basic-share dilution", "=Model!D32", "Review alongside fully diluted cap-table analysis.", "#,##0.0%"),
    (15, "Decision status", "=Model!D34", "Do not issue any grant if status is not READY FOR COMMITTEE REVIEW.", "General"),
]
for row, label, formula, interpretation, fmt in metrics:
    board[f"C{row}"] = label
    board[f"D{row}"] = formula
    board[f"D{row}"].font = Font(name="Arial", color=GREEN)
    board[f"D{row}"].number_format = fmt
    board[f"D{row}"].alignment = Alignment(horizontal="right")
    board[f"E{row}"] = interpretation

section(board, 19, "Committee checklist")
checks = [
    "1. Confirm latest FMV and cap-table capacity as-of the Board decision date.",
    "2. Confirm performance, role criticality and retention-risk selections have evidence and calibration notes.",
    "3. Confirm the grant is within the individual cap and the approved ESOP pool.",
    "4. Confirm legal documentation, vesting terms, strike price and tax treatment with counsel.",
    "5. Record the formal committee decision, conflicts of interest and grant communication owner.",
]
for idx, item in enumerate(checks, start=20):
    board[f"C{idx}"] = item
    board.merge_cells(start_row=idx, start_column=3, end_row=idx, end_column=8)

add_notes(board, 27, "This template does not replace a fully diluted cap-table, valuation opinion, tax review or Board resolution.")
board.print_area = "B2:H29"

for ws in [inputs, model, board]:
    for row in ws.iter_rows(min_row=8, max_row=ws.max_row, min_col=3, max_col=min(ws.max_column, 8)):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True, horizontal=cell.alignment.horizontal or "left")
    auto_fit(ws)

# Formatting corrections for title merge and body numeric widths
for ws in [inputs, model, board]:
    ws.sheet_properties.pageSetUpPr.fitToPage = True

wb.save(OUTPUT)
print(OUTPUT)
