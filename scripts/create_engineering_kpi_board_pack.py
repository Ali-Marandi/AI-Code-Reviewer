from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.comments import Comment
from openpyxl.formatting.rule import FormulaRule

OUTPUT = Path("/home/ubuntu/AI-Code-Reviewer/engineering_kpi_board_pack_template.xlsx")

DARK_GREEN = "135B44"
LIGHT_GREEN = "CFE9E0"
LIGHT_GRAY = "E7E5E4"
WHITE = "FFFFFF"
BLUE = "0000FF"
BLACK = "000000"
GREEN = "008000"
RED_FILL = "FCE4D6"
YELLOW_FILL = "FFF2CC"
GREEN_FILL = "E2F0D9"

thin_gray = Side(style="thin", color="A6A6A6")
medium_green = Side(style="medium", color=DARK_GREEN)


def base_tab(ws, title, subtitle, unit_note):
    ws["C3"] = title
    ws.merge_cells("C3:J3")
    ws["C3"].fill = PatternFill("solid", fgColor=DARK_GREEN)
    ws["C3"].font = Font(name="Arial", size=16, bold=True, color=WHITE)
    ws["C5"] = subtitle
    ws["C5"].font = Font(name="Arial", size=11, bold=True, color=BLACK)
    ws["C6"] = unit_note
    ws["C6"].font = Font(name="Arial", size=10, italic=True, color=BLACK)
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 20
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.oddFooter.center.text = "&A | Page &P"


def section(ws, row, title, end_col=10):
    ws.cell(row=row, column=3, value=title)
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=end_col)
    cell = ws.cell(row=row, column=3)
    cell.fill = PatternFill("solid", fgColor=LIGHT_GREEN)
    cell.font = Font(name="Arial", bold=True, color=BLACK)


def headers(ws, row, values):
    for col, value in enumerate(values, start=3):
        cell = ws.cell(row=row, column=col, value=value)
        cell.fill = PatternFill("solid", fgColor=LIGHT_GRAY)
        cell.font = Font(name="Arial", bold=True, color=BLACK)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=thin_gray)


def input_cell(ws, row, col, value=None, fmt="General", note="Source: Actual data from named system; period, extract filter and data owner must be documented."):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(name="Arial", color=BLUE)
    cell.number_format = fmt
    cell.alignment = Alignment(horizontal="right", vertical="top")
    cell.comment = Comment(note, "Manus AI")


def formula_cell(ws, row, col, formula, fmt="General", linked=False, bold=False):
    cell = ws.cell(row=row, column=col, value=formula)
    cell.font = Font(name="Arial", color=GREEN if linked else BLACK, bold=bold)
    cell.number_format = fmt
    cell.alignment = Alignment(horizontal="right", vertical="top")
    return cell


def autofit(ws):
    for col in range(3, ws.max_column + 1):
        letter = ws.cell(row=1, column=col).column_letter
        max_len = 10
        for row in range(1, ws.max_row + 1):
            value = ws.cell(row=row, column=col).value
            if value is not None:
                max_len = max(max_len, min(len(str(value)) + 2, 48))
        ws.column_dimensions[letter].width = max_len


def apply_common_alignment(ws):
    for row in ws.iter_rows(min_row=8, max_row=ws.max_row, min_col=3, max_col=min(ws.max_column, 10)):
        for cell in row:
            cell.alignment = Alignment(
                horizontal=cell.alignment.horizontal or "left",
                vertical="top",
                wrap_text=True,
            )


wb = Workbook()
inputs = wb.active
inputs.title = "Inputs"
dashboard = wb.create_sheet("KPI Dashboard")
board = wb.create_sheet("Board Pack")

# Inputs
base_tab(inputs, "Engineering KPI Board Pack — Inputs", "Blue-font cells are hardcoded operational inputs; targets require Board or management approval.", "(Values are intentionally blank; populate only with validated actual data.)")
section(inputs, 8, "Reporting control")
headers(inputs, 9, ["Field", "Value", "Source / definition", "Data owner"])
control_rows = [
    (10, "Reporting month", "e.g., 2026-08", "Board reporting calendar", "Board Ops"),
    (11, "Data cut-off date", None, "Timestamp after which data cannot change without restatement", "Data/Operations"),
    (12, "Data-quality reviewer", None, "Person accountable for completeness, freshness and reconciliation", "CTO delegate"),
    (13, "Baseline period", None, "First approved measurement period for trend comparison", "CTO + Board"),
]
for row, label, value, source, owner in control_rows:
    inputs.cell(row=row, column=3, value=label)
    input_cell(inputs, row, 4, value, "General", f"Source: Board reporting calendar / governed data record. {source}")
    inputs.cell(row=row, column=5, value=source)
    inputs.cell(row=row, column=6, value=owner)

section(inputs, 16, "KPI data-entry table")
headers(inputs, 17, ["KPI", "Direction", "Actual — prior month", "Actual — current month", "Board target", "Alert threshold", "Unit", "System of record"])

# KPI: name, direction, fmt, unit, source
kpis = [
    ("Regression pass rate", "Higher is better", "0.0%", "%", "CI / JUnit"),
    ("SAST coverage of in-scope paths", "Higher is better", "0.0%", "%", "Local SAST + CI/SARIF"),
    ("Confirmed high-severity security incidents", "Lower is better", "0", "count", "Incident register"),
    ("Open critical security-alert aging", "Lower is better", "0", "days", "Security backlog"),
    ("Lead time for change (median)", "Lower is better", "0.0", "hours/days", "GitHub + CI/CD"),
    ("Deployment frequency", "Higher is better", "0", "releases / month", "GitHub Releases + CI"),
    ("Failed deployment recovery time (median)", "Lower is better", "0.0", "hours", "Incident register + CI"),
    ("Change fail rate", "Lower is better", "0.0%", "%", "Release register + incidents"),
    ("Headcount vs. approved plan", "Higher is better", "0.0%", "%", "HRIS + hiring plan"),
    ("Time to hire (median)", "Lower is better", "0", "days", "ATS"),
    ("Regretted attrition (rolling 12 months)", "Lower is better", "0.0%", "%", "HRIS + exit interviews"),
    ("Pilot-customer health score", "Higher is better", "0.0%", "%", "CRM + product telemetry + support"),
    ("Key-feature adoption", "Higher is better", "0.0%", "%", "Opt-in product telemetry"),
    ("Bug-prediction precision", "Higher is better", "0.0%", "%", "Human-reviewed evaluation dataset"),
    ("Bug-prediction recall", "Higher is better", "0.0%", "%", "Curated ground-truth dataset"),
    ("Unactionable AI finding rate", "Lower is better", "0.0%", "%", "Review labels + customer feedback"),
]
for idx, (name, direction, fmt, unit, system) in enumerate(kpis, start=18):
    inputs.cell(row=idx, column=3, value=name)
    input_cell(inputs, idx, 4, direction, "General", "Source: KPI dictionary in ENGINEERING_BOARD_KPI_PACK.md. Direction changes require CTO approval.")
    input_cell(inputs, idx, 5, None, fmt, f"Source: {system}; prior full monthly period, using the approved KPI definition.")
    input_cell(inputs, idx, 6, None, fmt, f"Source: {system}; current reporting month through recorded data cut-off date.")
    input_cell(inputs, idx, 7, None, fmt, "Source: Board-approved KPI target; target must carry approval date and owner.")
    input_cell(inputs, idx, 8, None, fmt, "Source: Board-approved alert threshold; used for yellow/red status escalation.")
    inputs.cell(row=idx, column=9, value=unit)
    inputs.cell(row=idx, column=10, value=system)
inputs.freeze_panes = "C18"
inputs.print_area = "B2:J36"

# Dashboard
base_tab(dashboard, "Engineering KPI Dashboard", "Formula-based month-on-month dashboard; does not calculate a status until actuals and approved thresholds are entered.", "(Green-font values are linked from Inputs; status formulas are black.)")
section(dashboard, 8, "Scorecard")
headers(dashboard, 9, ["KPI", "Direction", "Prior month", "Current month", "Board target", "Alert threshold", "Variance vs. target", "Status"])
for output_row, source_row in enumerate(range(18, 18 + len(kpis)), start=10):
    formula_cell(dashboard, output_row, 3, f"=Inputs!C{source_row}", "General", linked=True)
    formula_cell(dashboard, output_row, 4, f"=Inputs!D{source_row}", "General", linked=True)
    fmt = kpis[source_row - 18][2]
    for col, input_col in zip([5, 6, 7, 8], [5, 6, 7, 8]):
        formula_cell(dashboard, output_row, col, f"=Inputs!{chr(64 + input_col)}{source_row}", fmt, linked=True)
    # variance: current - target; retains direction-independent raw difference, allowing audit
    formula_cell(dashboard, output_row, 9, f'=IF(OR(NOT(ISNUMBER(F{output_row})),NOT(ISNUMBER(G{output_row}))),"",F{output_row}-G{output_row})', fmt)
    # Green if current meets target; Yellow if only meets alert threshold; Red otherwise.
    formula_cell(dashboard, output_row, 10, f'=IF(OR(NOT(ISNUMBER(F{output_row})),NOT(ISNUMBER(G{output_row})),NOT(ISNUMBER(H{output_row}))),"DATA REQUIRED",IF(D{output_row}="Higher is better",IF(F{output_row}>=G{output_row},"GREEN",IF(F{output_row}>=H{output_row},"YELLOW","RED")),IF(F{output_row}<=G{output_row},"GREEN",IF(F{output_row}<=H{output_row},"YELLOW","RED"))))')

# Conditional formatting for status column
for label, fill in [("GREEN", GREEN_FILL), ("YELLOW", YELLOW_FILL), ("RED", RED_FILL)]:
    dashboard.conditional_formatting.add(
        f"J10:J{9 + len(kpis)}",
        FormulaRule(formula=[f'J10="{label}"'], fill=PatternFill("solid", fgColor=fill)),
    )

section(dashboard, 29, "Data governance note")
dashboard["C30"] = "A GREEN status only means the entered actual met the approved threshold; it is not proof that the underlying KPI is well-defined or that the system has no risk."
dashboard.merge_cells("C30:J30")
dashboard["C30"].font = Font(name="Arial", italic=True, color=BLACK)
dashboard["C31"] = "Review any YELLOW/RED KPI with a documented root cause, accountable owner, action date and Board decision request."
dashboard.merge_cells("C31:J31")
dashboard["C31"].font = Font(name="Arial", italic=True, color=BLACK)
dashboard.freeze_panes = "C10"
dashboard.print_area = "B2:J31"

# Board Pack
base_tab(board, "Monthly Engineering Board Pack", "Decision-oriented template for the AI-Code-Reviewer Enterprise 12-month engineering roadmap.", "(Populate narrative fields after dashboard data quality review.)")
section(board, 8, "A. Executive decision log")
headers(board, 9, ["Area", "Status", "Decision requested", "Accountable owner", "Decision due", "Risk if delayed"])
for row, area in enumerate(["Release quality", "Security / trust", "Velocity / operations", "Hiring / capability", "Pilot customer / product"], start=10):
    board.cell(row=row, column=3, value=area)
    for col in range(4, 9):
        input_cell(board, row, col, None, "General", "Source: Monthly Board Pack owner; state a specific decision, risk and date. Do not enter confidential raw customer data.")

section(board, 17, "B. KPI snapshot")
headers(board, 18, ["KPI", "Current month", "Board target", "Status", "Owner", "Root cause / 30-day action"])
# select 8 board-level KPI rows from dashboard with their source mapping
board_kpi_indices = [10, 11, 12, 14, 15, 18, 21, 23]
board_owner = ["QA Lead", "Security Engineer", "CTO + Security", "Engineering Lead", "DevOps Lead", "CTO + HR", "Product + CS", "AI Lead"]
for row, dash_row, owner in zip(range(19, 27), board_kpi_indices, board_owner):
    formula_cell(board, row, 3, f"='KPI Dashboard'!C{dash_row}", "General", linked=True)
    fmt = kpis[dash_row - 10][2]
    formula_cell(board, row, 4, f"='KPI Dashboard'!F{dash_row}", fmt, linked=True)
    formula_cell(board, row, 5, f"='KPI Dashboard'!G{dash_row}", fmt, linked=True)
    formula_cell(board, row, 6, f"='KPI Dashboard'!J{dash_row}", "General", linked=True)
    board.cell(row=row, column=7, value=owner)
    input_cell(board, row, 8, None, "General", "Source: KPI owner; include root cause, 30-day action, date and escalation need for any YELLOW/RED status.")

section(board, 29, "C. Roadmap milestones and risk escalation")
headers(board, 30, ["Workstream", "Milestone / evidence", "Status", "Risk / dependency", "Corrective action", "Owner"])
workstreams = ["Engine / AI", "Desktop / UX", "Security / Trust", "CI/CD / Operations", "Talent / Team"]
for row, workstream in enumerate(workstreams, start=31):
    board.cell(row=row, column=3, value=workstream)
    for col in range(4, 9):
        input_cell(board, row, col, None, "General", "Source: Workstream owner; cite verifiable release, PR, test report, approval or risk register entry.")

section(board, 38, "D. Meeting record")
headers(board, 39, ["Decision", "Owner", "Due date", "Evidence of completion", "Status", "Follow-up date"])
for row in range(40, 45):
    for col in range(3, 9):
        input_cell(board, row, col, None, "General", "Source: Board meeting minutes; capture final decision, accountable owner and verifiable evidence.")
board.print_area = "B2:J45"

for ws in [inputs, dashboard, board]:
    apply_common_alignment(ws)
    autofit(ws)

wb.save(OUTPUT)
print(OUTPUT)
