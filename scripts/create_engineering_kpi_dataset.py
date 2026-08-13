import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import csv
import os

def generate_datasets():
    output_dir = '/home/ubuntu/AI-Code-Reviewer/reports'
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    excel_path = os.path.join(output_dir, 'engineering_kpis_and_dora.xlsx')
    csv_path = os.path.join(output_dir, 'engineering_kpis_and_dora.csv')

    # 1. Create CSV dataset
    csv_data = [
        ["Metric Category", "Indicator Name", "Q1 Actual", "Q2 Actual", "Q3 Actual", "Q4 Projection", "Target / SLA", "Status"],
        ["DORA Metrics", "Deployment Frequency (Deploys/Month)", 5, 12, 18, 25, ">= 20", "On Track"],
        ["DORA Metrics", "Lead Time for Changes (Hours)", 12.0, 8.0, 4.0, 2.0, "< 4.0", "Exceeding"],
        ["DORA Metrics", "Mean Time to Recovery (MTTR - Hrs)", 2.5, 1.8, 1.0, 0.5, "< 1.0", "On Track"],
        ["DORA Metrics", "Change Failure Rate (%)", 4.5, 3.2, 2.0, 1.0, "< 2.0", "On Track"],
        ["Product Health", "SAST Finding Accuracy (%)", 90.0, 92.0, 94.0, 96.0, ">= 95.0", "Near Target"],
        ["Product Health", "AI Bug Prediction Precision (%)", 88.0, 91.0, 94.0, 95.5, ">= 90.0", "Exceeding"],
        ["Product Health", "Dependency Graph Analysis Pass (%)", 95.0, 96.5, 98.0, 99.0, ">= 98.0", "On Track"],
        ["Product Health", "Secret Scan False Positive Rate (%)", 3.5, 2.1, 1.2, 0.5, "< 1.0", "Near Target"],
        ["Team Scaling", "Engineering Headcount (Actual)", 4, 5, 8, 12, "12", "Aligned"],
        ["Team Scaling", "Engineering Headcount (Target)", 4, 6, 10, 14, "14", "Review"],
        ["Team Scaling", "Retention Rate - AI/Core (%)", 100.0, 100.0, 95.0, 95.0, ">= 90.0", "Exceeding"],
        ["Phase 1 Execution", "v1.1.0 Release Completion (%)", 0, 0, 100, 100, "100", "Completed"],
        ["Phase 1 Execution", "i18n Localization Foundation (%)", 0, 0, 100, 100, "100", "Completed"],
        ["Phase 1 Execution", "Security Audit & DPA Readiness (%)", 0, 50, 85, 100, "100", "In Progress"],
        ["Phase 1 Execution", "Design Partner Pilot Cohort (%)", 0, 30, 60, 100, "100", "In Progress"]
    ]

    with open(csv_path, mode='w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerows(csv_data)
    print(f"CSV dataset saved to: {csv_path}")

    # 2. Create Professional Excel Workbook with openpyxl
    wb = openpyxl.Workbook()

    # Sheet 1: Executive KPI Summary
    ws1 = wb.active
    ws1.title = "KPI & DORA Summary"
    ws1.views.sheetView[0].showGridLines = True

    # Colors & Styles
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    title_font = Font(name="Segoe UI", size=16, bold=True, color="0F172A")
    subtitle_font = Font(name="Segoe UI", size=11, italic=True, color="475569")
    data_font = Font(name="Segoe UI", size=10, color="000000")
    bold_data_font = Font(name="Segoe UI", size=10, bold=True, color="000000")
    border_thin = Border(left=Side(style='thin', color='CBD5E1'),
                         right=Side(style='thin', color='CBD5E1'),
                         top=Side(style='thin', color='CBD5E1'),
                         bottom=Side(style='thin', color='CBD5E1'))
    align_center = Alignment(horizontal='center', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')

    # Title Block
    ws1['A1'] = "AI-Code-Reviewer Enterprise — Engineering KPI & DORA Metrics Dashboard"
    ws1['A1'].font = title_font
    ws1['A2'] = "Comprehensive audit report and quantitative performance tracking for Board Review (August 2026)"
    ws1['A2'].font = subtitle_font

    ws1.append([]) # Row 3 blank

    # Headers (Row 4)
    headers = ["Metric Category", "Indicator Name", "Q1 Actual", "Q2 Actual", "Q3 Actual", "Q4 Projection", "Target / SLA", "Status"]
    ws1.append(headers)

    for col_num in range(1, len(headers) + 1):
        cell = ws1.cell(row=4, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center
        cell.border = border_thin

    ws1.row_dimensions[4].height = 25

    # Data Rows (Row 5 to 20)
    for row_idx, row_data in enumerate(csv_data[1:], start=5):
        ws1.append(row_data)
        ws1.row_dimensions[row_idx].height = 20
        for col_idx in range(1, len(row_data) + 1):
            cell = ws1.cell(row=row_idx, column=col_idx)
            cell.font = data_font
            cell.border = border_thin
            if col_idx in [1, 2]:
                cell.alignment = align_left
            elif col_idx in [3, 4, 5, 6, 7]:
                cell.alignment = align_right
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0.0' if isinstance(cell.value, float) else '#,##0'
            else:
                cell.alignment = align_center
                if cell.value in ["On Track", "Exceeding", "Completed"]:
                    cell.font = Font(name="Segoe UI", size=10, bold=True, color="047857") # Green
                elif cell.value in ["Near Target", "In Progress"]:
                    cell.font = Font(name="Segoe UI", size=10, bold=True, color="B45309") # Amber
                elif cell.value == "Review":
                    cell.font = Font(name="Segoe UI", size=10, bold=True, color="B91C1C") # Red

    # Auto-fit column widths
    for col in ws1.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws1.column_dimensions[col_letter].width = max(max_len + 5, 15)

    excel_path = os.path.join(output_dir, 'engineering_kpis_and_dora.xlsx')
    wb.save(excel_path)
    print(f"Excel dataset saved to: {excel_path}")
    return excel_path, csv_path

if __name__ == '__main__':
    generate_datasets()
